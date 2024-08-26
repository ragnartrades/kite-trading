from datetime import datetime, timedelta, date
import os
import re
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import List
from kiteconnect import KiteConnect
from common import utils, configs

kc: KiteConnect = utils.get_kite_connect_client()

# global variables, will be set after 'fresh_create_file_folders()' function is called
backtesting_directory = None
data_folder_path = None
# -----------------------------


class Input:
    START_DATE_TIME = "2024-08-12 09:15:00"
    END_DATE_TIME = "2024-08-12 15:30:00"
    # END_DATE_TIME = "2024-08-02 15:30:00"
    INTERVAL = 'minute'


def get_instrument_token_for_banknifty() -> str:
    global kc

    utils.update_nse_instruments(kc)
    nse_instruments_data: dict = utils.load_nse_instruments_data()

    return nse_instruments_data[configs.BANK_NIFTY_TRADING_SYMBOL]['instrument_token']


def strike_multiple_of_100(trading_symbol: str) -> bool:
    """ ex: trading_symbol = NIFTY24SEP24800CE """
    # Use a regular expression to find the integer before either 'CE' or 'PE'
    match = re.search(r'(\d+)(?:CE|PE)$', trading_symbol)
    if match:
        # Extract the integer
        number = int(match.group(1))
        # Check if it's a multiple of 100
        return number % 100 == 0
    return False


def should_be_in_requested_months_for_options(trading_symbol: str) -> bool:
    requested_months: List[str] = get_requested_months_from_option_date_range_input()
    expiry_month = get_expiry_month_for_option(trading_symbol)
    return expiry_month in requested_months


def get_requested_months_from_option_date_range_input() -> List[str]:
    start_date = datetime.strptime(Input.START_DATE_TIME,
                                   "%Y-%m-%d %H:%M:%S")
    end_date = datetime.strptime(Input.END_DATE_TIME,
                                 "%Y-%m-%d %H:%M:%S")

    distinct_months = set()

    current_date = start_date
    while current_date <= end_date:
        distinct_months.add(current_date.month)
        current_date += timedelta(days=1)

    month_symbols = []
    for month_number in distinct_months:
        month_symbols.append(utils.MONTH_NUMBER_TO_OPTION_MONTH_SYMBOL_MAP[month_number])

    return month_symbols


def get_expiry_month_for_option(trading_symbol: str) -> str:
    # Use a regular expression to find the text between two integers
    match = re.search(r'\d+(\D+)\d+', trading_symbol)
    if match:
        # Extract the text part between the numbers
        return match.group(1)

    return None


def get_instrument_tokens_map_for_options() -> dict:
    global kc

    utils.update_nfo_instruments(kc)
    nfo_instruments_data: dict = utils.load_nfo_instruments_data()

    instrument_tokens_map = {}
    for k, v in nfo_instruments_data.items():
        if k.startswith('BANKNIFTY') and \
                k.endswith(('CE', 'PE')) and \
                strike_multiple_of_100(k) and \
                should_be_in_requested_months_for_options(k):
            instrument_tokens_map[v['instrument_token']] = {
                'instrument_token': v['instrument_token'],
                'trading_symbol': v['tradingsymbol'],
            }

    return instrument_tokens_map


class BankNiftyDataInput(Input):
    INSTRUMENT_TOKEN = get_instrument_token_for_banknifty()


class OptionDataInput(Input):
    INSTRUMENT_TOKENS = get_instrument_tokens_map_for_options()


def GENERATE_CANDLESTICK_DATA_FOR_DATE_RANGE():
    fresh_create_file_folders()
    generate_banknifty_data_for_all_dates()
    generate_options_data_for_all_dates_for_all_options()


def fresh_create_file_folders():
    global backtesting_directory, data_folder_path

    print('deleting and recreating folders and files . . .')

    backtesting_directory = os.path.join(os.getcwd(), 'backtesting')

    # 1. Define the path to the data folder
    data_folder_path = os.path.join(backtesting_directory, 'data')

    # 2. Delete the existing data folder if it exists
    if os.path.exists(data_folder_path):
        shutil.rmtree(data_folder_path)

    # 3. Recreate the data folder
    os.makedirs(data_folder_path)

    # 4. create all sub folders an files
    dates: List[dict] = get_date_list_for_given_date_range(
        Input.START_DATE_TIME,
        Input.END_DATE_TIME,
    )

    for day in dates:
        # Set the current working directory to ./data
        os.chdir(data_folder_path)

        year = day['year']
        month = day['month']
        date = day['date']

        # 1. create year directory if required
        year_directory = os.path.join(data_folder_path, f'{year}')
        if not os.path.exists(year_directory):
            os.makedirs(year_directory)

        # 2. create month directory if required
        os.chdir(year_directory)
        month_directory = os.path.join(year_directory, f'{month}')
        if not os.path.exists(month_directory):
            os.makedirs(month_directory)

        # 3. create xlsx file
        file_path = os.path.join(month_directory, f'{date}.xlsx')
        wb = openpyxl.Workbook()
        wb.save(file_path)

    print('All folders and files are recreated !!!')


def get_date_list_for_given_date_range(
        start_date_time: str,
        end_date_time: str,
) -> List[dict]:
    my_start_date_time = datetime.strptime(start_date_time, "%Y-%m-%d %H:%M:%S")
    my_end_date_time = datetime.strptime(end_date_time, "%Y-%m-%d %H:%M:%S")

    # Generate a list of dictionaries for each date in the range
    date_list = []
    current_date_time = my_start_date_time

    while current_date_time.date() <= my_end_date_time.date():
        date_list.append({
            "year": current_date_time.year,
            "month": current_date_time.month,
            "date": current_date_time.day
        })

        current_date_time += timedelta(days=1)

    return date_list


def generate_banknifty_data_for_all_dates():
    my_start_date_time = datetime.strptime(BankNiftyDataInput.START_DATE_TIME,
                                           "%Y-%m-%d %H:%M:%S")
    my_end_date_time = datetime.strptime(BankNiftyDataInput.END_DATE_TIME,
                                         "%Y-%m-%d %H:%M:%S")

    current_date_time = my_start_date_time
    while current_date_time.date() <= my_end_date_time.date():
        generate_banknifty_data_for_specific_date(current_date_time.date())

        current_date_time += timedelta(days=1)


def generate_banknifty_data_for_specific_date(today: date):
    print(f'fetching banknifty candlestick data from kite. '
          f'instrument token: {BankNiftyDataInput.INSTRUMENT_TOKEN}')
    bank_nifty_candlestick_data = kc.historical_data(
        instrument_token=BankNiftyDataInput.INSTRUMENT_TOKEN,
        from_date=BankNiftyDataInput.START_DATE_TIME,
        to_date=BankNiftyDataInput.END_DATE_TIME,
        interval=BankNiftyDataInput.INTERVAL,
    )
    for candlestick in bank_nifty_candlestick_data:
        candlestick['date'] = candlestick['date'].strftime("%Y-%m-%d %H:%M:%S")

    target_file_path = get_target_file_path_from_date(today)
    sheet_name = 'BANKNIFTY'

    print(f'writing banknifty candle stick data to target file: {target_file_path}'
          f' and sheet: {sheet_name}')
    write_candlestick_data_to_target_file_and_sheet(bank_nifty_candlestick_data,
                                                    target_file_path, sheet_name)


def get_target_file_path_from_date(cur_date: date) -> str:
    year = cur_date.year
    month = cur_date.month
    day = cur_date.day

    target_file_path = f'{data_folder_path}/{year}/{month}/{day}.xlsx'

    return target_file_path


def write_candlestick_data_to_target_file_and_sheet(
        bank_nifty_candlestick_data: List[dict],
        target_file_path: str,
        sheet_name: str,
):
    df = pd.DataFrame(bank_nifty_candlestick_data)

    wb = openpyxl.load_workbook(target_file_path)

    # Check if the sheet name already exists, and if so, remove it
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)

    # Remove the default sheet if it exists
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    ws = wb.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(target_file_path)


def generate_options_data_for_all_dates_for_all_options():
    my_start_date_time = datetime.strptime(OptionDataInput.START_DATE_TIME,
                                           "%Y-%m-%d %H:%M:%S")
    my_end_date_time = datetime.strptime(OptionDataInput.END_DATE_TIME,
                                         "%Y-%m-%d %H:%M:%S")

    current_date_time = my_start_date_time
    while current_date_time.date() <= my_end_date_time.date():
        generate_options_data_for_specific_date_for_all_options(current_date_time.date())

        current_date_time += timedelta(days=1)


def generate_options_data_for_specific_date_for_all_options(today: date):
    for option_instrument_token, option_instrument_data in OptionDataInput.INSTRUMENT_TOKENS.items():
        generate_options_data_for_specific_date_for_specific_option(
            today,
            option_instrument_data,
        )


def generate_options_data_for_specific_date_for_specific_option(
        today: date,
        option_instrument_data: dict,
):
    print(f'fetching banknifty option candlestick data from kite. '
          f'instrument token: {option_instrument_data['instrument_token']}, '
          f'trading symbol: {option_instrument_data['trading_symbol']}')

    bank_nifty_option_candlestick_data = kc.historical_data(
        instrument_token=option_instrument_data['instrument_token'],
        from_date=BankNiftyDataInput.START_DATE_TIME,
        to_date=BankNiftyDataInput.END_DATE_TIME,
        interval=BankNiftyDataInput.INTERVAL,
    )
    for candlestick in bank_nifty_option_candlestick_data:
        candlestick['date'] = candlestick['date'].strftime("%Y-%m-%d %H:%M:%S")

    target_file_path = get_target_file_path_from_date(today)
    sheet_name = option_instrument_data['trading_symbol']

    print(f'writing banknifty option candle stick data to target file: {target_file_path}'
          f' and sheet: {sheet_name}')

    write_candlestick_data_to_target_file_and_sheet(bank_nifty_option_candlestick_data,
                                                    target_file_path, sheet_name)


if __name__ == '__main__':
    GENERATE_CANDLESTICK_DATA_FOR_DATE_RANGE()
