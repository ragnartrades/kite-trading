# This is gap opening STBT long straddle backtesting over BANKNIFTY monthly
# expiry for AUGUST month

from datetime import datetime, date, timedelta
from typing import TypedDict, List
from openpyxl import load_workbook
import pandas as pd

from common.utils import MONTH_NUMBER_TO_OPTION_MONTH_SYMBOL_MAP

source_data_dir = '/Users/Subhram/my-projects/kite-trading/backtesting/data'
report_file_path = ('/Users/Subhram/my-projects/kite-trading/backtesting/'
                    'gap_opening_long_straddle_report.xlsx')


class Input:
    start_date = date(2024, 9, 1)
    end_date = date(2024, 9, 9)
    stock = 'BANKNIFTY'


class Report(TypedDict):
    stock: str

    entry_date: date
    entry_stock_point: float
    entry_ce_price: float
    entry_pe_price: float
    entry_tot_price: float

    strike: float

    exit_date: date
    exit_stock_point: float
    exit_ce_price: float
    exit_pe_price: float
    exit_tot_price: float

    summary_stock_point_delta: float
    summary_ce_delta: float
    summary_pe_delta: float
    summary_tot_delta: float
    summary_basic_profit_percentage: float


def get_total_fund_required_considering_BTST_margin(report: Report) -> float:
    fund_during_buying = report['entry_tot_price'] * 1.2
    extra_margin_during_selling = report['exit_tot_price'] * 0.2

    return fund_during_buying + extra_margin_during_selling


def report_to_dict(report: Report) -> dict:
    summary_ce_delta = report['exit_ce_price'] - report['entry_ce_price']
    summary_pe_delta = report['exit_pe_price'] - report['entry_pe_price']
    summary_tot_delta = summary_ce_delta + summary_pe_delta
    tot_fund_required = get_total_fund_required_considering_BTST_margin(report)
    summary_basic_profit_percentage = (summary_tot_delta / tot_fund_required) * 100

    return {
        'stock': report['stock'],

        'entry_date': report['entry_date'].strftime("%Y-%m-%d"),
        'entry_stock_point': report['entry_stock_point'],
        'entry_ce_price': report['entry_ce_price'],
        'entry_pe_price': report['entry_pe_price'],
        'entry_tot_price': report['entry_tot_price'],

        'strike': report['strike'],

        'exit_date': report['exit_date'].strftime("%Y-%m-%d"),
        'exit_stock_point': report['exit_stock_point'],
        'exit_ce_price': report['exit_ce_price'],
        'exit_pe_price': report['exit_pe_price'],
        'exit_tot_price': report['exit_tot_price'],

        'summary_stock_point_delta': report['exit_stock_point'] - report['entry_stock_point'],
        'summary_ce_delta': summary_ce_delta,
        'summary_pe_delta': summary_pe_delta,
        'summary_tot_delta': summary_tot_delta,
        'summary_basic_profit_percentage': summary_basic_profit_percentage,
    }


class AllReports:
    reports: List[Report] = []

    @classmethod
    def _clean_worksheet(cls, ws):
        row = 3
        while row <= 100:
            for cell in ws[row]:
                cell.value = None

            row += 1

    @classmethod
    def append_to_report_sheet(cls):
        print('writing report ...')

        wb = load_workbook(report_file_path)
        ws = wb.active

        cls._clean_worksheet(ws)

        # If the workbook is empty or needs initialization, create a new sheet
        if ws.title != 'report':
            ws.title = 'report'

        # If reports are not empty, append rows to the worksheet
        if cls.reports:
            df = pd.DataFrame([report_to_dict(r) for r in cls.reports])

            # Define the starting row
            start_row = 3

            # Append each row to the worksheet starting from the specified row
            for i, row in df.iterrows():
                for j, value in enumerate(row):
                    # Write data to the worksheet
                    ws.cell(row=start_row + i, column=j + 1, value=value)

        # Save and close the file
        wb.save(report_file_path)
        wb.close()

        print('report writing finished !!!')


def get_source_file_path_from_source_dir_and_date(source_data_dir: str, cur_date: date):
    year = cur_date.year
    month = cur_date.month
    day = cur_date.day

    return f'{source_data_dir}/{year}/{month}/{day}.xlsx'


def get_stock_eod_price(cur_date: date) -> float:
    source_data_file_path = get_source_file_path_from_source_dir_and_date(
        source_data_dir, cur_date)

    last_close = None
    df = pd.read_excel(source_data_file_path, sheet_name='BANKNIFTY')
    for row_num, data in df.iterrows():
        close: float = data['close']
        last_close = close

    return last_close


def get_suitable_strike_price_from_stock_price(entry_stock_price: float) -> int:
    lower_strike_price = (entry_stock_price // 100) * 100
    # return lower_strike_price

    upper_strike_price = lower_strike_price + 100

    if abs(upper_strike_price - entry_stock_price) < abs(lower_strike_price - entry_stock_price):
        return int(upper_strike_price)

    return int(lower_strike_price)


def get_option_instrument_symbol_from_strike_and_date(strike: int, cur_date: date, option_type: str) \
        -> str:
    year = cur_date.year
    last_two_digit_in_year = year % 2000

    month_num = cur_date.month
    month_str = MONTH_NUMBER_TO_OPTION_MONTH_SYMBOL_MAP[month_num]

    return f'BANKNIFTY{last_two_digit_in_year}{month_str}{int(strike)}{option_type}'


def get_option_eod_price(option_symbol: str, cur_date: date) -> float:
    source_data_file_path = get_source_file_path_from_source_dir_and_date(
        source_data_dir, cur_date)

    last_close = None
    df = pd.read_excel(source_data_file_path, sheet_name=option_symbol)
    for row_num, data in df.iterrows():
        close: float = data['close']
        last_close = close

    return last_close


def get_sod_stock_price(cur_date: date) -> float:
    source_data_file_path = get_source_file_path_from_source_dir_and_date(
        source_data_dir, cur_date)

    df = pd.read_excel(source_data_file_path, sheet_name='BANKNIFTY')
    for row_num, data in df.iterrows():
        return data['open']


def get_option_sod_price(option_symbol: str, cur_date: date) -> float:
    source_data_file_path = get_source_file_path_from_source_dir_and_date(
        source_data_dir, cur_date)

    df = pd.read_excel(source_data_file_path, sheet_name=option_symbol)
    for row_num, data in df.iterrows():
        return data['open']


def not_a_trading_day(cur_date: date) -> bool:
    source_data_file_path = get_source_file_path_from_source_dir_and_date(
        source_data_dir, cur_date)

    df = pd.read_excel(source_data_file_path, sheet_name='BANKNIFTY')
    for row_num, data in df.iterrows():
        return False

    return True


def last_date_for_testing(cur_date: date) -> bool:
    return cur_date == Input.end_date


def back_test_for_cur_date(cur_date: date) -> Report:
    next_date = cur_date + timedelta(days=1)

    if last_date_for_testing(cur_date) or \
            not_a_trading_day(cur_date) or not_a_trading_day(next_date):
        return None

    print(f'starting backtesting for date: {cur_date.__str__()} ...')

    entry_stock_price = get_stock_eod_price(cur_date)

    strike = get_suitable_strike_price_from_stock_price(entry_stock_price)
    ce_symbol = get_option_instrument_symbol_from_strike_and_date(strike, cur_date, 'CE')
    pe_symbol = get_option_instrument_symbol_from_strike_and_date(strike, cur_date, 'PE')

    ce_entry_price = get_option_eod_price(ce_symbol, cur_date)
    pe_entry_price = get_option_eod_price(pe_symbol, cur_date)
    entry_tot_price = ce_entry_price + pe_entry_price

    if not_a_trading_day(next_date):
        return None

    exit_stock_point = get_sod_stock_price(next_date)

    ce_exit_price = get_option_sod_price(ce_symbol, next_date)
    pe_exit_price = get_option_sod_price(pe_symbol, next_date)
    exit_tot_price = ce_exit_price + pe_exit_price

    ce_delta = ce_exit_price - ce_entry_price
    pe_delta = pe_exit_price - pe_entry_price
    tot_delta = ce_delta + pe_delta
    basic_profit_percentage = (tot_delta / entry_tot_price) * 100

    print(f'backtesting finished for date: {cur_date.__str__()} !!!')

    return Report(
        stock=Input.stock,
        entry_date=cur_date,
        entry_stock_point=entry_stock_price,
        entry_ce_price=ce_entry_price,
        entry_pe_price=pe_entry_price,
        entry_tot_price=entry_tot_price,
        strike=strike,
        exit_date=next_date,
        exit_stock_point=exit_stock_point,
        exit_ce_price=ce_exit_price,
        exit_pe_price=pe_exit_price,
        exit_tot_price=exit_tot_price,
        summary_stock_point_delta=exit_stock_point - entry_stock_price,
        summary_ce_delta=ce_delta,
        summary_pe_delta=pe_delta,
        summary_tot_delta=tot_delta,
        summary_basic_profit_percentage=basic_profit_percentage,
    )


def back_test():
    cur_date = Input.start_date
    while cur_date <= Input.end_date:
        report: Report = back_test_for_cur_date(cur_date)
        if report is not None:
            AllReports.reports.append(report)

        cur_date += timedelta(days=1)

    AllReports.append_to_report_sheet()


if __name__ == '__main__':
    back_test()
